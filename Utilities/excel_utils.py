import openpyxl

def access_products(excel_path, sheet_name):
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook[sheet_name]
    start_row = 2
    product_names = []

    for row in sheet.iter_rows(start_row, values_only=True):
        product_name = row[0]
        product_names.append(product_name)

    return product_names

def access_prices(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    prices = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        price = str(row[0])
        prices.append(price)
    return prices